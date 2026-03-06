"""
Export search results to Excel.

One sheet per search term.
Columns: Title | Authors | Abstract | DOI | Link | Source(s)
Optional extra columns: Sentiment | Confidence
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from app.config import EXCEL_COLUMNS, EXCEL_SENTIMENT_COLUMNS

# ── Colour palette (hex, no #) ────────────────────────────────────────────────
_HEADER_BG = "1C1C1E"
_HEADER_FG = "FFFFFF"
_ACCENT = "0A84FF"
_ROW_ALT = "2C2C2E"
_ROW_MAIN = "1C1C1E"
_TEXT = "EBEBF5"
_MUTED = "8E8E93"

_THIN_SIDE = Side(style="thin", color="38383A")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)


def _header_style(cell):
    cell.font = Font(bold=True, color=_HEADER_FG, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=_ACCENT)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _THIN_BORDER


def _data_style(cell, alt_row: bool = False):
    bg = _ROW_ALT if alt_row else _ROW_MAIN
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(color=_TEXT, name="Calibri", size=10)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = _THIN_BORDER


def _col_widths(has_sentiment: bool) -> list[int]:
    # Title, Authors, Abstract, DOI, Link, Source(s) [, Sentiment, Confidence]
    widths = [40, 30, 70, 25, 35, 20]
    if has_sentiment:
        widths += [14, 14]
    return widths


def _sheet_name(term: str) -> str:
    # Excel sheet names max 31 chars, no special chars
    safe = "".join(c if c.isalnum() or c in " _-" else "_" for c in term)
    return safe[:31]


_EXCEL_MAX_CHARS = 32767  # Excel hard limit per cell


def _safe_str(value) -> str:
    """Convert a value to string and truncate to Excel's per-cell character limit."""
    s = str(value) if value is not None else ""
    return s[:_EXCEL_MAX_CHARS]


def export_all(papers: list[dict], output_dir: str) -> str:
    """
    Write all papers to a single 'All Results' sheet.
    Papers must have 'matched_queries': list[str] from the global dedup step.
    Returns the full path of the saved file.
    """
    out_path = Path(output_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    if not os.access(str(out_path), os.W_OK):
        raise PermissionError(
            f"Export folder is not writable: {output_dir}\n"
            "Please choose a different folder."
        )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"PaperScraper_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    columns = ["Title", "Authors", "Abstract", "DOI", "Link", "Source(s)", "Matched Queries"]
    widths = [40, 30, 70, 25, 35, 20, 50]

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws = wb.create_sheet(title="All Results")
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = _ACCENT

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _header_style(cell)

    for row_idx, paper in enumerate(papers, start=2):
        alt = row_idx % 2 == 0
        matched = paper.get("matched_queries", [])
        row_data = [
            _safe_str(paper.get("title", "")),
            _safe_str(paper.get("authors", "")),
            _safe_str(paper.get("abstract", "")),
            _safe_str(paper.get("doi", "")),
            _safe_str(paper.get("url", "")),
            _safe_str(paper.get("source", "")),
            _safe_str("; ".join(matched)),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _data_style(cell, alt)
            col_name = columns[col_idx - 1]
            if col_name == "DOI" and value:
                doi_url = value if value.startswith("http") else f"https://doi.org/{value}"
                cell.hyperlink = doi_url
                cell.font = Font(color=_ACCENT, underline="single", name="Calibri", size=10)
            if col_name == "Link" and value:
                cell.hyperlink = value
                cell.font = Font(color=_ACCENT, underline="single", name="Calibri", size=10)

    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(filepath)
    return filepath


def export(
    results_by_term: dict[str, list[dict]],
    output_dir: str,
    include_sentiment: bool = False,
) -> str:
    """
    Write one Excel workbook to output_dir.
    results_by_term: {search_term: [paper_dict, ...]}
    Returns the full path of the saved file.
    Raises PermissionError if output_dir is not writable.
    """
    out_path = Path(output_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    if not os.access(str(out_path), os.W_OK):
        raise PermissionError(
            f"Export folder is not writable: {output_dir}\n"
            "Please choose a different folder."
        )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"PaperScraper_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    # Remove default blank sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    columns = EXCEL_SENTIMENT_COLUMNS if include_sentiment else EXCEL_COLUMNS

    for term, papers in results_by_term.items():
        ws = wb.create_sheet(title=_sheet_name(term))

        # ── Header row ────────────────────────────────────────────────────────
        ws.row_dimensions[1].height = 28
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            _header_style(cell)

        ws.freeze_panes = "A2"

        # ── Data rows ─────────────────────────────────────────────────────────
        for row_idx, paper in enumerate(papers, start=2):
            alt = row_idx % 2 == 0
            row_data = [
                _safe_str(paper.get("title", "")),
                _safe_str(paper.get("authors", "")),
                _safe_str(paper.get("abstract", "")),
                _safe_str(paper.get("doi", "")),
                _safe_str(paper.get("url", "")),
                _safe_str(paper.get("source", "")),
            ]
            if include_sentiment:
                row_data += [
                    _safe_str(paper.get("sentiment", "")),
                    paper.get("confidence", ""),  # float — no truncation needed
                ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                _data_style(cell, alt)
                # Make DOI and URL clickable
                col_name = columns[col_idx - 1]
                if col_name == "DOI" and value:
                    doi_url = value if value.startswith("http") else f"https://doi.org/{value}"
                    cell.hyperlink = doi_url
                    cell.font = Font(color=_ACCENT, underline="single",
                                     name="Calibri", size=10)
                if col_name == "Link" and value:
                    cell.hyperlink = value
                    cell.font = Font(color=_ACCENT, underline="single",
                                     name="Calibri", size=10)

        # ── Column widths ─────────────────────────────────────────────────────
        for col_idx, width in enumerate(_col_widths(include_sentiment), start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── Summary row at top (row 0 space / just a note) ───────────────────
        ws.sheet_properties.tabColor = _ACCENT

    wb.save(filepath)
    return filepath
