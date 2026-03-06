"""
Sentiment Analysis tab.

Allows the user to:
  1. Load papers from the current Search session OR import an Excel file.
  2. Enter a topic to score sentiment against.
  3. Optionally supply a HuggingFace API token.
  4. Run analysis (with per-paper progress bar).
  5. Browse results with Sentiment + Confidence columns.
  6. Export the enriched results to Excel.
"""

from __future__ import annotations

import threading
from pathlib import Path
from tkinter import filedialog
from typing import Callable, Optional

import customtkinter as ctk

from app.config import COLORS, FONT_FAMILY
from app.analysis.sentiment import analyse_papers
from app.export.excel_exporter import export
from app.storage import history as db


class SentimentTab(ctk.CTkFrame):

    def __init__(self, parent, get_papers_fn: Callable[[], list[dict]]):
        super().__init__(parent, fg_color=COLORS["bg_primary"], corner_radius=0)

        self._get_papers_fn = get_papers_fn
        self._papers: list[dict] = []
        self._results: list[dict] = []
        self._running = False
        self._output_dir: str = str(Path.home() / "Desktop")

        saved_dir = db.load_setting("output_dir")
        if saved_dir and Path(saved_dir).exists():
            self._output_dir = saved_dir

        self._build_layout()

    # ── Layout ────────────────────────────────────────────────────────────────

    def _build_layout(self):
        self.columnconfigure(0, weight=0, minsize=290)
        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)

        sidebar = ctk.CTkFrame(self, fg_color=COLORS["bg_secondary"], corner_radius=0, width=290)
        sidebar.grid(row=0, column=0, sticky="nsew")
        sidebar.grid_propagate(False)

        main = ctk.CTkFrame(self, fg_color=COLORS["bg_primary"], corner_radius=0)
        main.grid(row=0, column=1, sticky="nsew")

        self._build_sidebar(sidebar)
        self._build_main(main)

    def _build_sidebar(self, sb):
        pad = {"padx": 16, "pady": (10, 4)}

        # ── Data source ───────────────────────────────────────────────────────
        ctk.CTkLabel(
            sb, text="Data Source",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            text_color=COLORS["text_primary"], anchor="w",
        ).pack(fill="x", **pad)

        ctk.CTkButton(
            sb, text="Use Current Search Results", height=38,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["bg_input"],
            text_color=COLORS["text_secondary"],
            corner_radius=8,
            command=self._load_from_search,
        ).pack(fill="x", padx=16, pady=(0, 6))

        ctk.CTkButton(
            sb, text="Import from Excel File…", height=38,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["bg_input"],
            text_color=COLORS["text_secondary"],
            corner_radius=8,
            command=self._import_excel,
        ).pack(fill="x", padx=16, pady=(0, 4))

        self._source_label = ctk.CTkLabel(
            sb, text="No papers loaded.",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=COLORS["text_muted"], anchor="w",
        )
        self._source_label.pack(fill="x", padx=16, pady=(0, 8))

        _sep(sb)

        # ── Topic input ───────────────────────────────────────────────────────
        ctk.CTkLabel(
            sb, text="Analysis Topic",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            text_color=COLORS["text_primary"], anchor="w",
        ).pack(fill="x", **pad)

        ctk.CTkLabel(
            sb,
            text="What topic should the sentiment be scored against?",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=COLORS["text_muted"],
            anchor="w", justify="left", wraplength=258,
        ).pack(fill="x", padx=16, pady=(0, 6))

        self._topic_entry = ctk.CTkEntry(
            sb,
            placeholder_text="e.g. CRISPR gene editing safety",
            fg_color=COLORS["bg_input"],
            border_color=COLORS["separator"],
            text_color=COLORS["text_primary"],
            placeholder_text_color=COLORS["text_muted"],
            corner_radius=8,
            height=36,
        )
        self._topic_entry.pack(fill="x", padx=16, pady=(0, 4))

        # Restore saved topic
        saved_topic = db.load_setting("sentiment_topic", "")
        if saved_topic:
            self._topic_entry.insert(0, saved_topic)

        self._topic_entry.bind(
            "<FocusOut>",
            lambda _: db.save_setting("sentiment_topic", self._topic_entry.get().strip()),
        )

        _sep(sb)

        # ── API token ─────────────────────────────────────────────────────────
        ctk.CTkLabel(
            sb, text="HuggingFace API Token",
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            text_color=COLORS["text_primary"], anchor="w",
        ).pack(fill="x", **pad)

        ctk.CTkLabel(
            sb,
            text="Optional — increases rate limits on the free API.\nNot saved to disk for security.",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=COLORS["text_muted"],
            anchor="w", justify="left", wraplength=258,
        ).pack(fill="x", padx=16, pady=(0, 6))

        self._token_entry = ctk.CTkEntry(
            sb,
            placeholder_text="hf_…",
            show="•",
            fg_color=COLORS["bg_input"],
            border_color=COLORS["separator"],
            text_color=COLORS["text_primary"],
            placeholder_text_color=COLORS["text_muted"],
            corner_radius=8,
            height=36,
        )
        self._token_entry.pack(fill="x", padx=16, pady=(0, 4))
        # Token is intentionally NOT persisted to disk — enter each session

        _sep(sb)

        # ── Export folder ─────────────────────────────────────────────────────
        _lbl(sb, "Export folder")
        dir_row = ctk.CTkFrame(sb, fg_color="transparent")
        dir_row.pack(fill="x", padx=16, pady=(0, 8))
        dir_row.columnconfigure(0, weight=1)
        self._dir_label = ctk.CTkLabel(
            dir_row,
            text=_short_path(self._output_dir),
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=COLORS["text_muted"],
            anchor="w",
        )
        self._dir_label.grid(row=0, column=0, sticky="ew")
        ctk.CTkButton(
            dir_row, text="Browse", width=64, height=28,
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["bg_input"],
            text_color=COLORS["text_secondary"],
            corner_radius=6,
            command=self._choose_dir,
        ).grid(row=0, column=1, padx=(8, 0))

        _sep(sb)

        # ── Analyse button ────────────────────────────────────────────────────
        self._analyse_btn = ctk.CTkButton(
            sb, text="Run Sentiment Analysis", height=44,
            font=ctk.CTkFont(family=FONT_FAMILY, size=14, weight="bold"),
            fg_color=COLORS["accent"],
            hover_color=COLORS["accent_hover"],
            corner_radius=10,
            command=self._start_analysis,
        )
        self._analyse_btn.pack(fill="x", padx=16, pady=(8, 4))

        # Progress
        self._progress_bar = ctk.CTkProgressBar(
            sb, height=6,
            fg_color=COLORS["bg_tertiary"],
            progress_color=COLORS["accent"],
            corner_radius=3,
        )
        self._progress_bar.set(0)
        self._progress_bar.pack(fill="x", padx=16, pady=(4, 2))

        self._progress_label = ctk.CTkLabel(
            sb, text="",
            font=ctk.CTkFont(family=FONT_FAMILY, size=11),
            text_color=COLORS["text_muted"], anchor="w",
        )
        self._progress_label.pack(fill="x", padx=16)

    def _build_main(self, m):
        m.rowconfigure(1, weight=1)
        m.columnconfigure(0, weight=1)

        # ── Column headers ────────────────────────────────────────────────────
        header = ctk.CTkFrame(m, fg_color=COLORS["bg_tertiary"], height=36, corner_radius=0)
        header.grid(row=0, column=0, sticky="ew", padx=0, pady=(12, 0))
        header.grid_propagate(False)

        cols = [
            ("Title", 260), ("Authors", 150), ("Abstract", 310),
            ("Sentiment", 100), ("Confidence", 100),
        ]
        for i, (name, width) in enumerate(cols):
            ctk.CTkLabel(
                header, text=name, width=width,
                font=ctk.CTkFont(family=FONT_FAMILY, size=11, weight="bold"),
                text_color=COLORS["text_muted"],
                anchor="w",
            ).pack(side="left", padx=(12 if i == 0 else 4, 0))

        # ── Results table ─────────────────────────────────────────────────────
        self._table_scroll = ctk.CTkScrollableFrame(
            m, fg_color=COLORS["bg_secondary"], corner_radius=12,
            scrollbar_button_color=COLORS["bg_tertiary"],
            scrollbar_button_hover_color=COLORS["accent"],
        )
        self._table_scroll.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 6))
        self._table_scroll.columnconfigure(0, weight=1)

        self._placeholder = ctk.CTkLabel(
            self._table_scroll,
            text="Load papers and run analysis to see results here.",
            font=ctk.CTkFont(family=FONT_FAMILY, size=14),
            text_color=COLORS["text_muted"],
        )
        self._placeholder.pack(pady=60)

        # ── Bottom bar ────────────────────────────────────────────────────────
        bottom = ctk.CTkFrame(m, fg_color=COLORS["bg_secondary"], corner_radius=12, height=50)
        bottom.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 12))
        bottom.grid_propagate(False)
        bottom.columnconfigure(0, weight=1)

        self._stats_label = ctk.CTkLabel(
            bottom, text="",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=COLORS["text_muted"], anchor="w",
        )
        self._stats_label.grid(row=0, column=0, sticky="w", padx=16, pady=12)

        self._export_btn = ctk.CTkButton(
            bottom, text="Export to Excel", width=160, height=34,
            font=ctk.CTkFont(family=FONT_FAMILY, size=13, weight="bold"),
            fg_color=COLORS["success"],
            hover_color="#28b84e",
            text_color="#000000",
            corner_radius=8,
            command=self._export,
            state="disabled",
        )
        self._export_btn.grid(row=0, column=1, padx=(0, 16), pady=8)

    # ── Data loading ───────────────────────────────────────────────────────────

    def _load_from_search(self):
        papers = self._get_papers_fn()
        if not papers:
            self._source_label.configure(
                text="No search results yet. Run a search first.",
                text_color=COLORS["warning"],
            )
            return
        self._papers = papers
        self._source_label.configure(
            text=f"{len(papers)} papers loaded from search.",
            text_color=COLORS["success"],
        )

    def _import_excel(self):
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls")],
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            papers = []
            for ws in wb.worksheets:
                headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
                col_map = {h: i for i, h in enumerate(headers)}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    def _get(key):
                        idx = col_map.get(key, -1)
                        return str(row[idx] or "") if idx >= 0 else ""
                    papers.append({
                        "title": _get("Title"),
                        "authors": _get("Authors"),
                        "abstract": _get("Abstract"),
                        "doi": _get("DOI"),
                        "url": _get("Link"),
                        "source": _get("Source(s)"),
                    })
            self._papers = papers
            self._source_label.configure(
                text=f"{len(papers)} papers loaded from {Path(path).name}.",
                text_color=COLORS["success"],
            )
        except Exception as exc:
            self._source_label.configure(
                text=f"Import error: {exc}",
                text_color=COLORS["error"],
            )

    # ── Analysis ───────────────────────────────────────────────────────────────

    def _start_analysis(self):
        if self._running:
            return
        if not self._papers:
            self._source_label.configure(
                text="Load papers first.", text_color=COLORS["warning"]
            )
            return
        topic = self._topic_entry.get().strip()
        if not topic:
            self._topic_entry.configure(border_color=COLORS["error"])
            self.after(800, lambda: self._topic_entry.configure(border_color=COLORS["separator"]))
            return

        self._running = True
        self._results.clear()
        self._analyse_btn.configure(text="Analysing…", state="disabled")
        self._export_btn.configure(state="disabled")
        self._progress_bar.set(0)
        self._clear_table()

        token = self._token_entry.get().strip() or None
        total = len(self._papers)

        def _cb(done: int, tot: int):
            def _update():
                self._progress_bar.set(done / tot)
                self._progress_label.configure(text=f"{done} / {tot} papers analysed")
            self.after(0, _update)

        def _worker():
            results = analyse_papers(self._papers, topic, api_token=token, progress_callback=_cb)
            self._results = results

            def _finish():
                self._running = False
                self._analyse_btn.configure(text="Run Sentiment Analysis", state="normal")
                self._render_results(results)
                self._stats_label.configure(
                    text=f'{len(results)} papers analysed for topic: "{topic}"',
                    text_color=COLORS["text_secondary"],
                )
                self._export_btn.configure(state="normal")

            self.after(0, _finish)

        threading.Thread(target=_worker, daemon=True).start()

    # ── Table rendering ────────────────────────────────────────────────────────

    def _clear_table(self):
        for w in self._table_scroll.winfo_children():
            w.destroy()

    def _render_results(self, results: list[dict]):
        self._clear_table()
        if not results:
            ctk.CTkLabel(
                self._table_scroll, text="No results.",
                font=ctk.CTkFont(family=FONT_FAMILY, size=14),
                text_color=COLORS["text_muted"],
            ).pack(pady=60)
            return

        for idx, paper in enumerate(results):
            alt = idx % 2 == 0
            bg = COLORS["bg_tertiary"] if alt else COLORS["bg_secondary"]

            sentiment = paper.get("sentiment", "error")
            confidence = paper.get("confidence", 0.0)

            sentiment_color = {
                "positive": COLORS["success"],
                "negative": COLORS["error"],
                "neutral": COLORS["text_muted"],
            }.get(sentiment, COLORS["warning"])

            row = ctk.CTkFrame(
                self._table_scroll, fg_color=bg, corner_radius=6, cursor="hand2",
            )
            row.pack(fill="x", padx=4, pady=2)

            ctk.CTkLabel(
                row, text=_trunc(paper.get("title", ""), 55),
                width=256, anchor="w", justify="left", wraplength=248,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
                text_color=COLORS["text_primary"],
            ).pack(side="left", padx=(10, 4), pady=8)

            ctk.CTkLabel(
                row, text=_trunc(paper.get("authors", ""), 35),
                width=146, anchor="w",
                font=ctk.CTkFont(family=FONT_FAMILY, size=11),
                text_color=COLORS["text_muted"],
            ).pack(side="left", padx=4, pady=8)

            ctk.CTkLabel(
                row, text=_trunc(paper.get("abstract", ""), 120),
                anchor="w", justify="left", wraplength=310,
                font=ctk.CTkFont(family=FONT_FAMILY, size=11),
                text_color=COLORS["text_secondary"],
            ).pack(side="left", padx=4, pady=8, expand=True, fill="x")

            ctk.CTkLabel(
                row, text=sentiment.capitalize(), width=96,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12, weight="bold"),
                text_color=sentiment_color, anchor="center",
            ).pack(side="left", padx=4, pady=8)

            conf_pct = f"{confidence * 100:.1f}%"
            ctk.CTkLabel(
                row, text=conf_pct, width=96,
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                text_color=COLORS["text_muted"], anchor="center",
            ).pack(side="left", padx=(4, 10), pady=8)

            for widget in row.winfo_children():
                widget.bind("<Button-1>", lambda e, p=paper: self._show_detail(p))
            row.bind("<Button-1>", lambda e, p=paper: self._show_detail(p))

    def _show_detail(self, paper: dict):
        win = ctk.CTkToplevel(self)
        win.title(paper.get("title", "Paper Details")[:60])
        win.geometry("700x580")
        win.configure(fg_color=COLORS["bg_primary"])
        win.grab_set()

        scroll = ctk.CTkScrollableFrame(win, fg_color=COLORS["bg_primary"])
        scroll.pack(fill="both", expand=True, padx=20, pady=20)

        def _field(label, value, color=None):
            ctk.CTkLabel(
                scroll, text=label,
                font=ctk.CTkFont(family=FONT_FAMILY, size=11, weight="bold"),
                text_color=COLORS["text_muted"], anchor="w",
            ).pack(fill="x", pady=(10, 2))
            ctk.CTkLabel(
                scroll, text=value or "—",
                font=ctk.CTkFont(family=FONT_FAMILY, size=12),
                text_color=color or COLORS["text_primary"],
                anchor="w", justify="left", wraplength=640,
            ).pack(fill="x")

        _field("Title", paper.get("title"))
        _field("Authors", paper.get("authors"))
        _field("Abstract", paper.get("abstract"))

        sentiment = paper.get("sentiment", "error")
        s_color = {"positive": COLORS["success"], "negative": COLORS["error"],
                   "neutral": COLORS["text_muted"]}.get(sentiment, COLORS["warning"])
        _field("Sentiment", f"{sentiment.capitalize()} ({paper.get('confidence', 0)*100:.1f}% confidence)", s_color)
        _field("Source(s)", paper.get("source"))

        ctk.CTkButton(
            win, text="Close", width=100,
            fg_color=COLORS["bg_tertiary"],
            hover_color=COLORS["bg_input"],
            text_color=COLORS["text_secondary"],
            command=win.destroy,
        ).pack(pady=(0, 16))

    # ── Export ─────────────────────────────────────────────────────────────────

    def _export(self):
        if not self._results:
            return
        topic = self._topic_entry.get().strip() or "sentiment"
        results_map = {f"Sentiment — {topic}": self._results}
        try:
            path = export(results_map, self._output_dir, include_sentiment=True)
            self._stats_label.configure(
                text=f"Exported to: {path}", text_color=COLORS["success"]
            )
        except Exception as exc:
            self._stats_label.configure(
                text=f"Export failed: {exc}", text_color=COLORS["error"]
            )

    # ── Settings ───────────────────────────────────────────────────────────────

    def _choose_dir(self):
        d = filedialog.askdirectory(title="Select export folder", initialdir=self._output_dir)
        if d:
            self._output_dir = d
            db.save_setting("output_dir", d)
            self._dir_label.configure(text=_short_path(d))


# ── Helpers ───────────────────────────────────────────────────────────────────

def _trunc(text: str, n: int) -> str:
    text = str(text).strip()
    return text[:n] + "…" if len(text) > n else text


def _sep(parent):
    ctk.CTkFrame(parent, fg_color=COLORS["separator"], height=1, corner_radius=0).pack(
        fill="x", padx=12, pady=6
    )


def _lbl(parent, text: str):
    ctk.CTkLabel(
        parent, text=text,
        font=ctk.CTkFont(family=FONT_FAMILY, size=11),
        text_color=COLORS["text_muted"], anchor="w",
    ).pack(fill="x", padx=16, pady=(4, 2))


def _short_path(path: str) -> str:
    p = Path(path)
    try:
        rel = p.relative_to(Path.home())
        return f"~/{rel}"
    except ValueError:
        parts = p.parts
        return str(Path(*parts[-2:])) if len(parts) > 2 else path
